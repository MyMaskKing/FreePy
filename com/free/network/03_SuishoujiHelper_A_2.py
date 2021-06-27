# coding=UTF-8
'''
03_SuishoujiHelper_A_2说明：随手记自动化工具所使用的流水数据文件做成
Created on 2020年4月2日

@author: dapao
'''
import os
import sys
import re
import csv
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import requests
import xlsxwriter
from xlutils.copy import copy
import datetime
from com.free.utils import MailUtils
from com.free.utils import GetNumByImg
# 用于解决爬取的数据格式化
import io
from future.backports.misc import count
from _ast import Try
from test.support import catch_threading_exception
"""
※重要：共通方法的导入
"""
from com.free.constant import CommonConstants
from com.free.utils.LogUtils import getDebugObj, getErrorObj
mydebug = getDebugObj()
myerror = getErrorObj()

####################################################
## 00_1路径及系统变量
####################################################
work_file_folder_path = r"C:/myfree_config/freePy/network/SuishoujiHelper/流水数据文件/"
result_file_folder_path = r"C:/myfree_config/freePy/network/SuishoujiHelper/作业报告/"
rule_file_path = r"C:/myfree_config/freePy/network/SuishoujiHelper/作业报告/第一步_随手记自动记账规则制定.xlsx"
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding=CommonConstants.ENCODING_UTF8)
sys.setrecursionlimit(1000000)
####################################################
## 00_2共同变量
####################################################
rule_fenlei_zhichu_list = []
rule_fenlei_shoru_list = []
rule_zhanghu_list = []
#####################################################
### 00_3:Excel 样式
###
######################################################
excelDom = xlsxwriter.Workbook(result_file_folder_path + "第二步_记账数据_待确认(自动记账用).xlsx");
format_disable = excelDom.add_format({
    "fg_color": "#969486","font_color": "#A6A6A6"  # 字体颜色
    })
format_excepton = excelDom.add_format({'bold': True, 'font_color': 'red',"fg_color": "yellow",})
format1 = excelDom.add_format({
    'font_color': '#963634','bold': True,"fg_color": "#B7DEE8"
})
####################################################
## 1,将支付宝和微信的流水文件合并成可以被随手记自动化利用的文件
####################################################
def crt_suishoujiInputFile():
    mydebug.logger.debug("开始读取工作CSV文件")
    # Excel作成命令开启
    sheel1Dom = excelDom.add_worksheet("随手记登录流水数据(支付宝和微信)")
    sheel1Dom.set_column("A:G", 25)
    sheel1Dom.set_column("H:H", 35)
    # Excel Title写入
    excel_title = ["收入/支出/转账","分类","分类ID","账户","账户ID","金额","时间","备注","原始数据"]
    excel_row_data_write_format(0,sheel1Dom,excel_title,format1)
    default_row_index_Val = 1
    # 读取对象文件夹下所有的子文件夹及文件
    #for root,dirs,files in os.walk(work_file_folder_path,topdown=False):
    listDir=os.listdir(work_file_folder_path)
        # 读取当前目录下的所有子目录
        #for dir in dirs:
            #print (os.path.join(root,dir))
        # 读取当前目录下的所有文件
    for file in listDir:
        if ".csv" in file :
            file_path = work_file_folder_path + file
            mydebug.logger.debug("工作文件" + file_path)
            default_row_index_Val = all_file_data_create(file_path,excelDom,sheel1Dom,default_row_index_Val)
    # 流水CSV文件内容合并后保存成Excel
    excelDom.close()

####################################################
## 1-1,支付宝和微信开始读取
####################################################
def readRuleExcel():
    global rule_fenlei_zhichu_list
    global rule_fenlei_shoru_list
    global rule_zhanghu_list
    mydebug.logger.debug("开始读取【随手记自动记账规则制定】工作EXCEL文件")
    wb = openpyxl.load_workbook(rule_file_path)
    sheets = wb.sheetnames
    print(sheets[0])
    sheetDom = wb[sheets[0]]
    sheet_max_row = sheetDom.max_row
    sheet_max_col = sheetDom.max_column
    print('【随手记自动记账规则制定】工作EXCEL文件：{}行 {}列'.format(sheet_max_row, sheet_max_col))
    title_flg = 0
    rule_fenlei_zhichu_flg = 0
    rule_fenlei_shouru_flg = 0
    rule_zhanghu_flg = 0
    for row_index in range(1,sheet_max_row+1): # 行
        tmp_row_data =[]
        for col_index in range(1,sheet_max_col+1): #列
            cell_val = sheetDom.cell(row_index,col_index).value
            if cell_val == "（随手记网站）分类名（支出）":
                rule_fenlei_shouru_flg = 0
                rule_zhanghu_flg = 0
                rule_fenlei_zhichu_flg = 1
                title_flg = 1
                break
            elif cell_val == "（随手记网站）分类名（收入）":
                rule_fenlei_zhichu_flg = 0
                rule_zhanghu_flg = 0
                rule_fenlei_shouru_flg = 1
                title_flg = 1
                break
            elif cell_val == "（随手记网站）账户名":
                rule_fenlei_zhichu_flg = 0
                rule_fenlei_shouru_flg = 0
                rule_zhanghu_flg = 1
                title_flg = 1
                break
            if sheetDom.cell(row_index,1).value == None :
                title_flg = 1
                break
            if cell_val != None :
                tmp_row_data.append(cell_val)
        if title_flg == 1:
            title_flg = 0
            continue
        if rule_fenlei_zhichu_flg == 1 :
            rule_fenlei_zhichu_list.append(tmp_row_data)
        if rule_fenlei_shouru_flg == 1 :
            rule_fenlei_shoru_list.append(tmp_row_data)
        if rule_zhanghu_flg == 1 :
            rule_zhanghu_list.append(tmp_row_data)
    print(len(rule_fenlei_zhichu_list),len(rule_fenlei_shoru_list),len(rule_zhanghu_list))
    return rule_fenlei_zhichu_list,rule_fenlei_shoru_list,rule_zhanghu_list

####################################################
## 1-2,支付宝和微信开始读取
####################################################
def all_file_data_create(file_path,excelDom,sheel1Dom,index_Val):
    rule_fenlei_zhichu_list,rule_fenlei_shoru_list,rule_zhanghu_list = readRuleExcel()
    work_file_nm = ""
    ######################
    # 文件类型：1支付宝 2微信 #
    ######################
    work_file_typ = 0
    default_row_index_Val = index_Val
    defined_encode = ""
    if "alipay_record_" in file_path :
        defined_encode = "gbk"
        work_file_typ = 1
    else:
        defined_encode = "utf-8"
        work_file_typ = 2
    # 开始读取支付/微信的CSV文件
    mydebug.logger.debug("文件解码Code：%s"%defined_encode)
    with open(file_path,encoding=defined_encode) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        # 判断当前行是否为CSV的Title
        csv_file_index = 1
        for row_data in csv_reader:
            # 流水文件的单元格总数大于10的话为流水数据
            if len(row_data) > 10 :
                # 当前行为CSV的Title时不读取
                if csv_file_index == 1 :
                    csv_file_index = 0
                    continue
                mydebug.logger.debug(row_data)
                tmp_row_data = []
                # Excel数据写入
                # 用法 add_sheet.write(行,列,label=写入内容)
                if work_file_typ == 1 :
                    # 支付宝流水文件处理
                    # 交易类型 Matching
                    if "其他"  in row_data[0]:
                        if '还款成功' in row_data[6]:
                            tmp_row_data.append("转账")
                        elif '交易成功' in row_data[6] and ('卖出' in row_data[3] or '买入' in row_data[3]) :
                            tmp_row_data.append("转账")
                        elif ('退款成功' in row_data[6]) :
                            tmp_row_data.append("收入")
                        elif '交易成功' in row_data[6] and '收益发放' in row_data[3] :
                            tmp_row_data.append("收入")
                        elif '已关闭' in row_data[6] :
                            continue
                        else:
                            tmp_row_data.append("(异常)未知交易类型")
                    else:
                        tmp_row_data.append(row_data[0])
                    # Matching_账户: 文件名,收/付款方式,交易状态 
                    match_zhanghu_row_data = has_flg_by_list(rule_zhanghu_list,[file_path,row_data[4],row_data[6]])
                    # Matching_分类(支出): 文件名,收/付款方式,交易状态 
                    match_fenlei_zhichu_row_data = has_flg_by_list(rule_fenlei_zhichu_list,[file_path,row_data[3],row_data[6],row_data[7]])
                    # Matching_分类(收入): 文件名,收/付款方式,交易状态 
                    match_fenlei_shoru_row_data = has_flg_by_list(rule_fenlei_shoru_list,[file_path,row_data[3],row_data[6],row_data[7]])
                    if match_fenlei_zhichu_row_data != None :
                        tmp_row_data.append(match_fenlei_zhichu_row_data[0])
                        tmp_row_data.append(match_fenlei_zhichu_row_data[1])
                    elif match_fenlei_shoru_row_data != None :
                        tmp_row_data.append(match_fenlei_shoru_row_data[0])
                        tmp_row_data.append(match_fenlei_shoru_row_data[1])
                    else:
                        tmp_row_data.append("(异常)请填入分类名(收入)")
                        tmp_row_data.append("(异常)请填入分类ID(收入)")
                    if match_zhanghu_row_data != None :
                        tmp_row_data.append(match_zhanghu_row_data[0])
                        tmp_row_data.append(match_zhanghu_row_data[1])
                    else:
                        tmp_row_data.append("(异常)请填入账户名")
                        tmp_row_data.append("(异常)请填入账户ID")
                    
                    # 金额
                    tmp_row_data.append(row_data[5])
                    # 交易时间
                    tmp_row_data.append(row_data[10].replace("/","."))
                    # 备注
                    tmp_row_data.append(row_data[3]+",交易分类 ："+row_data[7])
                    # 原始数据
                    tmp_row_data.append(','.join(row_data))
                    excel_row_data_write(default_row_index_Val,sheel1Dom,tmp_row_data)
                else :
                    # 微信流水文件处理
                    excel_row_data_write_by_weixin(default_row_index_Val,sheel1Dom,row_data)
                default_row_index_Val = default_row_index_Val + 1
    return default_row_index_Val

####################################################
## 99-0:共通方法：Excel的行数据写入(共同)
####################################################
def has_flg_by_list(match_list,target_list):
    for match_data in match_list:
        if len(match_data) > 2:
            match_str = match_data[2]
            # 含有多个匹配值
            if ("," in match_str) :
                match_success_count = 1
                match_str_1 = match_str.split(",")
                for ms1 in match_str_1:
                    neq_flg = 0
                    if "!" in ms1 :
                        ms1 = ms1.replace("!","")
                        neq_flg = 1
                    for target_data in target_list:
                        if ms1 in target_data :
                            match_success_count = match_success_count +1
                            if neq_flg == 1:
                                neq_flg = 0
                                match_success_count = 0
                                break
                if match_success_count == len(match_str_1) :
                    return match_data
            # 1个匹配值
            else:
                for target_data in target_list:
                    if match_str in target_data :
                        return match_data
####################################################
## 99-0:共通方法：Excel的行数据写入(共同)
####################################################
def excel_row_data_write(rowNum, sheetDom, dataArray):
    indexVal = 0
    for data in dataArray:
        if indexVal == 8 :
            sheetDom.write(rowNum,indexVal,data,format_disable);
        elif "(异常)" in dataArray[indexVal]:
            sheetDom.write(rowNum,indexVal,data,format_excepton);
        else:
            sheetDom.write(rowNum,indexVal,data);
        indexVal = indexVal + 1
####################################################
## 99-1:共通方法：Excel的行数据写入(支付宝)
####################################################
def excel_row_data_write_by_zhifubao(rowNum, excelDom, dataArray):
    excelDom.write(rowNum,0,label = dataArray[0]);
    excelDom.write(rowNum,1,label = dataArray[3]);
    excelDom.write(rowNum,2,label = dataArray[4]);
    excelDom.write(rowNum,3,label = dataArray[5]);
    excelDom.write(rowNum,4,label = dataArray[10]);
    excelDom.write(rowNum,5,label = dataArray[1]);
####################################################
## 99-2:共通方法：Excel的行数据写入(微信)
####################################################
def excel_row_data_write_by_weixin(rowNum, excelDom, dataArray):
    excelDom.write(rowNum,0,label = dataArray[4]);
    excelDom.write(rowNum,1,label = dataArray[3]);
    excelDom.write(rowNum,2,label = dataArray[6]);
    excelDom.write(rowNum,3,label = dataArray[0]);
    excelDom.write(rowNum,3,label = dataArray[0]);
####################################################
## 99-3:共通方法：Excel的行数据写入(共同)样式
####################################################
def excel_row_data_write_format(rowNum, sheetDom, dataArray, format):
    indexVal = 0
    for data in dataArray:
        sheetDom.write(rowNum,indexVal,data, format);
        indexVal = indexVal + 1

if __name__ == '__main__':
    crt_suishoujiInputFile()
#     MailUtils.sendmail("GetNetData", "xinxi")
#     networObj = adjustExist(url)
#     workbook = xlrd.open_workbook(file_path + r"\Py1.xlsx");
#     copy = copy(workbook)
#     copy.save(file_path + r"\Py1.xlsx")
